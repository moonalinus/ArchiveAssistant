import os
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import fitz
import re
from collections import defaultdict
import pytesseract
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog

# Путь к исполняемому файлу Tesseract OCR
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Словарь для перевода номеров месяцев в русские названия
months = {
    1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля', 5: 'мая', 6: 'июня',
    7: 'июля', 8: 'августа', 9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'
}

def extract_info(pdf_file):
    try:
        # Откройте PDF-файл с помощью PyMuPDF
        with fitz.open(pdf_file) as pdf:
            # Получите содержимое первой страницы в виде текста
            page = pdf[0]
            text = page.get_text()

            # Найдите номер выпуска в году
            issue_number_in_year = re.search(r'№\s*(\d+)', text)
            issue_number_in_year = issue_number_in_year.group(1) if issue_number_in_year else None

            # Найдите номер выпуска среди всех выпусков
            total_issue_number = re.search(r'\((\d+)\)', text)
            total_issue_number = total_issue_number.group(1) if total_issue_number else None

            # Найдите дату выпуска
            date = re.search(r'\d+\s+\w+\s+\d+\s*года?', text)
            issue_date = date.group() if date else None

            # Проверьте наличие надписи "специальный выпуск"
            special_issue = re.search(r'специальный\s+выпуск', text, re.IGNORECASE)
            special_issue_text = special_issue.group() if special_issue else None

            return issue_number_in_year, total_issue_number, issue_date, special_issue_text
    except Exception as e:
        print(f"Ошибка при обработке файла {pdf_file}: {e}")
        return None, None, None, None

def extract_page_number(pdf_file):
    try:
        # Откройте PDF-файл с помощью PyMuPDF
        with fitz.open(pdf_file) as pdf:
            # Получите содержимое первой страницы в виде текста
            page = pdf[0]
            text = page.get_text()

            # Найдите номер страницы
            page_number = re.search(r'\b(1|2|3|4|5|6|7|8|9|10|11|12)\b', text, re.IGNORECASE)
            page_number = page_number.group() if page_number else None

            return page_number
    except Exception as e:
        print(f"Ошибка при обработке файла {pdf_file}: {e}")
        return None

def process_folder(root_folder, progress_bar):
    # Создайте новую книгу Excel
    workbook = openpyxl.Workbook()

    # Создайте новый лист в книге Excel
    worksheet = workbook.active

    # Запишите заголовки столбцов
    headers = ['Общий порядковый номер файла', 'Номер электронного документа по описи', 'Текст', 'Дата изменения', 'Размер (байты)', 'Формат файла']
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Получите список всех папок внутри текущей папки, отсортированный по имени
    subfolders = [f.path for f in os.scandir(root_folder) if f.is_dir()]
    subfolders.sort()

    # Создайте словарь для хранения данных о файлах, отсортированных по дате изменения
    date_time = defaultdict(list)

    # Инициализируйте общий порядковый номер файла
    overall_file_num = 1

    # Получите список всех PDF файлов в всех подпапках
    all_pdf_files = []
    for subfolder in subfolders:
        files = [f.path for f in os.scandir(subfolder) if f.is_file() and f.name.endswith('.pdf')]
        all_pdf_files.extend(files)

    # Инициализируйте счетчик обработанных файлов
    file_count = 0

    # Обработайте каждую папку
    for subfolder in subfolders:
        # Получите список файлов в текущей папке, отсортированный по имени
        files = [f.path for f in os.scandir(subfolder) if f.is_file() and f.name.endswith('.pdf')]
        files.sort()
# Обработайте первый файл в папке
        if files:
            first_file = files[0]
            issue_number_in_year, total_issue_number, issue_date, special_issue_text = extract_info(first_file)

            # Обработайте каждый файл в папке
            for file_num, file_path in enumerate(files, 1):
                file_stats = os.stat(file_path)
                file_size = file_stats.st_size
                file_modified = file_stats.st_mtime
                file_format = os.path.splitext(file_path)[1].upper().replace(".","")

                # Преобразуйте дату последнего изменения в нужный формат
                file_modified_date = datetime.datetime.fromtimestamp(file_modified).date()
                file_modified_str = f"{file_modified_date.day} {months[file_modified_date.month]} {file_modified_date.year}"

                # Добавьте данные о файле в словарь, отсортированный по дате изменения
                page_num = extract_page_number(file_path)
                file_info = (overall_file_num, "1", f"{page_num} страница газеты «Сталинградская трибуна»\n№ {issue_number_in_year} ({total_issue_number}) от {issue_date}", file_modified_str, file_size, file_format)
                date_time[file_modified_date].append(file_info)

                # Увеличьте общий порядковый номер файла
                overall_file_num += 1

                # Увеличьте счетчик обработанных файлов
                file_count += 1

                # Обновите шкалу прогресса
                progress_bar.step(100 / len(all_pdf_files))

    # Запишите данные из отсортированного словаря в таблицу Excel
    row_num = 2
    for date, file_data in sorted(date_time.items()):
        for file_info in file_data:
            worksheet.cell(row=row_num, column=1, value=file_info[0])
            worksheet.cell(row=row_num, column=2, value=file_info[1])
            worksheet.cell(row=row_num, column=3, value=file_info[2])
            worksheet.cell(row=row_num, column=4, value=file_info[3])
            worksheet.cell(row=row_num, column=5, value=file_info[4])
            worksheet.cell(row=row_num, column=6, value=file_info[5])
            row_num += 1

    # Автоматически настройте ширину столбцов
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].auto_size = True

    # Сохраните книгу Excel
    workbook.save('magazine.xlsx')

    # Обновите шкалу прогресса
    progress_bar.step(100 - progress_bar['value'])

    # Выведите сообщение об удачном завершении обработки
    print("Обработка завершена успешно!")

# Создайте графический интерфейс
root = tk.Tk()
root.title("PDF Processing")

# Создайте поле ввода для пути к папке
root_folder_var = tk.StringVar()
root_folder_entry = ttk.Entry(root, width=50, textvariable=root_folder_var)
root_folder_entry.pack(pady=10)

# Создайте кнопку для выбора папки
browse_button = ttk.Button(root, text="Browse...", command=lambda: select_folder())
browse_button.pack(pady=10)

# Функция для выбора папки
def select_folder():
    folder_path = filedialog.askdirectory(initialdir=os.getcwd(), title="Select folder")
    root_folder_var.set(folder_path)

# Создайте шкалу прогресса
progress_bar = ttk.Progressbar(root, length=200, mode='determinate')
progress_bar.pack(pady=10)

# Создайте кнопку для запуска обработки
start_button = ttk.Button(root, text="Start processing", command=lambda: process_folder(root_folder_var.get(), progress_bar))
start_button.pack(pady=10)

# Запустите главный цикл приложения
root.mainloop()