import os
import re
import openpyxl
import datetime
from PIL import Image

def get_page_number_word(page_number):
    page_numbers = {
        "1": "Первая",
        "2": "Вторая",
        "3": "Третья",
        "4": "Четвертая",
        "5": "Пятая",
        "6": "Шестая",
        "7": "Седьмая",
        "8": "Восьмая",
        "9": "Девятая",
        "10": "Десятая"
    }
    return page_numbers.get(page_number, page_number)

def get_image_modification_date(file_path):
    with Image.open(file_path) as img:
        return img._getexif().get(36867)

def format_date(date_string):
    date = datetime.datetime.strptime(date_string, '%Y:%m:%d %H:%M:%S')
    months = {
        1: 'января',
        2: 'февраля',
        3: 'марта',
        4: 'апреля',
        5: 'мая',
        6: 'июня',
        7: 'июля',
        8: 'августа',
        9: 'сентября',
        10: 'октября',
        11: 'ноября',
        12: 'декабря'
    }
    return f"{date.day} {months[date.month]} {date.year}"

def process_files(folder_path):
    # Создаем новую книгу Excel и получаем активный лист
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Получаем список папок и сортируем их по номеру
    folders = sorted(os.listdir(folder_path), key=lambda x: int(re.findall(r'\d+', x)[-1]))

    # Инициализируем переменную для порядкового номера файла
    file_number = 1

    # Проходимся по каждой папке
    for folder in folders:
        folder_full_path = os.path.join(folder_path, folder)

        # Получаем список файлов в папке и сортируем их по номеру газеты и номеру страницы
        files = sorted(os.listdir(folder_full_path), key=lambda x: int(x.split(" - ")[0]))

        # Проходимся по каждому файлу
        for file in files:
            file_full_path = os.path.join(folder_full_path, file)

            # Извлекаем информацию из названия файла
            match = re.match(r'(\d+) - №(\d+)-(\d+)', file)
            newspaper_number = match.group(2)
            page_number = match.group(3)

            # Получаем информацию о файле
            file_extension = os.path.splitext(file)[1]
            file_size_mb = round(os.path.getsize(file_full_path) / (1024 * 1024), 2)

            # Получаем номер страницы буквами
            page_number_word = get_page_number_word(page_number)

            # Формируем название газеты
            # Тут добавить код, чтобы можно было менять начальное чило общего номера газет
            newspaper_name = f"{page_number_word} страница газеты «Еланский колхозник»\n№ {newspaper_number} ({int(newspaper_number) + 4217}) от "

            # Получаем дату последнего изменения файла
            mod_date = None
            mod_date_timestamp = os.path.getmtime(file_full_path)
            mod_date = datetime.datetime.fromtimestamp(mod_date_timestamp).strftime('%Y:%m:%d %H:%M:%S')
            mod_date = format_date(mod_date)

            # Записываем информацию в таблицу Excel
            sheet.cell(row=file_number, column=1, value=file_number)
            sheet.cell(row=file_number, column=2, value=newspaper_name)
            sheet.cell(row=file_number, column=3, value="")
            sheet.cell(row=file_number, column=4, value=file_size_mb)
            sheet.cell(row=file_number, column=5, value=mod_date)

            file_number += 1

    # Сохраняем книгу Excel
    workbook.save("output.xlsx")

# Пример использования
folder_path = input("Введите путь к папке с файлами: ")
process_files(folder_path)


# /home/specialist/Рабочий стол/1953 Еланский колхозник - № с 81 по 88 м
# /home/specialist/Рабочий стол/Елань описи и папки/Елань 1957 обработано
