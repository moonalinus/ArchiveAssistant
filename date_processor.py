import openpyxl
from openpyxl import load_workbook
from datetime import datetime, timedelta
import re
import locale

# Устанавливаем локаль для русского языка
locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')

# Функция для преобразования строки с датой в формат datetime
def parse_date(date_string):
    months = {
        'января': 1, 'февраля': 2, 'марта': 3, 'апреля': 4, 'мая': 5, 'июня': 6,
        'июля': 7, 'августа': 8, 'сентября': 9, 'октября': 10, 'ноября': 11, 'декабря': 12
    }
    date_pattern = r'(\d{1,2}) (\w+) (\d{4}) года'
    match = re.search(date_pattern, date_string)
    if match:
        day, month, year = match.groups()
        return datetime(int(year), months[month], int(day))
    return None

# Функция для преобразования даты в формат строки с русскими названиями месяцев
def format_date(date):
    day = date.strftime('%d')
    month = date.strftime('%B')
    year = date.strftime('%Y')
    if int(day) < 10:
        day = day.lstrip('0')
    return f'{day} {month} 1957 года'# {year} года'

# Функция для обработки Excel файла
def process_excel(file_name):
    # Загружаем рабочую книгу
    wb = load_workbook(filename=file_name)
    # Обрабатываем каждый лист в книге
    for sheet in wb.worksheets:
        # Обрабатываем каждую ячейку в листе
        for row in sheet.iter_rows():
            for cell in row:
                # Пытаемся преобразовать значение ячейки в дату
                date = parse_date(cell.value)
                if date:
                    # Вычитаем один день и преобразуем дату обратно в строку
                    new_date = format_date(date - timedelta(days=1))
                    # Записываем новое значение в ячейку
                    cell.value = new_date
    # Сохраняем изменения в файл
    wb.save(file_name)


# Вызываем функцию process_excel, передавая имя файла в качестве аргумента
process_excel('date_in.xlsx')